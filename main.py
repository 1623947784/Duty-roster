from __future__ import annotations

import argparse
import re
import sys
from copy import copy
from dataclasses import dataclass
from math import ceil
from pathlib import Path
from typing import Iterable, List

import pandas as pd
from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.utils import column_index_from_string


SCHEDULE_FILE = Path('江苏大学2024-2025学年第2学期课程表.xlsx')
TEMPLATE_FILE = Path('附表2 江苏大学实习、课程设计教学日历模板.xlsx')
OUTPUT_FILE = TEMPLATE_FILE.with_name(f"{TEMPLATE_FILE.stem}_填充.xlsx")


WEEKDAY_ORDER = {
    '星期一': 1,
    '星期二': 2,
    '星期三': 3,
    '星期四': 4,
    '星期五': 5,
}


@dataclass
class Session:
    week_number: int
    week_label: str
    weekday: str
    period_label: str
    course_name: str
    content: str
    teacher: str

    @property
    def description(self) -> str:
        lines: List[str] = []
        if self.course_name:
            lines.append(self.course_name)
        if self.content:
            content_text = self.content.replace('\r', '\n')
            for segment in re.split(r'[；;\n]+', content_text):
                normalized = segment.strip()
                if normalized:
                    lines.append(normalized)
        if self.teacher:
            lines.append(self.teacher)
        return '\n'.join(lines)


def extract_week_number(raw_value) -> int | None:
    if pd.isna(raw_value):
        return None
    text = str(raw_value).strip()
    if not text or text.lower() == 'nan':
        return None
    normalized = text.replace('第', '').replace('周', '')
    if not re.fullmatch(r'\d+(\.0)?', normalized):
        return None
    try:
        return int(float(normalized))
    except ValueError:
        return None


def normalize_cell(value) -> str:
    if pd.isna(value):
        return ''
    text = str(value).strip()
    if not text or text.lower() == 'nan':
        return ''
    return text.replace('\r\n', '\n').replace('\r', '\n')


def split_cell_payload(text: str) -> tuple[str, str, str]:
    if not text:
        return '', '', ''
    parts = [segment.strip() for segment in text.split('\n') if segment.strip()]
    if not parts:
        return '', '', ''
    course = parts[0]
    if len(parts) == 1:
        return course, '', ''
    teacher = parts[-1]
    content = '；'.join(parts[1:-1]) if len(parts) > 2 else ''
    return course, content, teacher


def build_sessions(schedule_path: Path) -> List[Session]:
    if not schedule_path.exists():
        raise FileNotFoundError(f'未找到课程表文件：{schedule_path}')

    df = pd.read_excel(schedule_path, header=[1, 2, 3])
    period_row = df.iloc[0]
    data_rows = df.iloc[1:]

    week_col = df.columns[0]

    sessions: List[Session] = []

    for _, row in data_rows.iterrows():
        week_number = extract_week_number(row[week_col])
        if week_number is None:
            continue
        week_label = f'第{week_number}周'

        for col in df.columns[2:]:
            weekday = col[1]
            if weekday not in WEEKDAY_ORDER:
                continue

            cell_text = normalize_cell(row[col])
            if not cell_text:
                continue

            course_name, content, teacher = split_cell_payload(cell_text)
            if not any([course_name, content, teacher]):
                continue

            period_value = period_row[col]
            period_label = ''
            if not pd.isna(period_value):
                try:
                    period_label = str(int(float(period_value)))
                except (TypeError, ValueError):
                    period_label = normalize_cell(period_value)

            sessions.append(
                Session(
                    week_number=week_number,
                    week_label=week_label,
                    weekday=weekday,
                    period_label=period_label,
                    course_name=course_name,
                    content=content,
                    teacher=teacher,
                )
            )

    sessions.sort(
        key=lambda s: (
            s.week_number,
            WEEKDAY_ORDER.get(s.weekday, 99),
            int(s.period_label) if s.period_label.isdigit() else 99,
            s.course_name,
        )
    )

    return sessions


def write_sessions_to_template(template_path: Path, sessions: Iterable[Session], output_path: Path) -> None:
    if not template_path.exists():
        raise FileNotFoundError(f'未找到模板文件：{template_path}')

    sessions = list(sessions)
    wb = load_workbook(template_path)
    ws = wb.active

    header_cell = next(
        (cell for cell in ws['A'] if isinstance(cell.value, str) and cell.value.strip() == '日程安排'),
        None,
    )
    if header_cell is None:
        raise ValueError('未在模板中找到“日程安排”标题。')
    data_start_row = header_cell.row + 1

    footer_cell = next(
        (
            cell
            for cell in ws['A']
            if isinstance(cell.value, str) and cell.value.strip().startswith('考核方式')
        ),
        None,
    )
    if footer_cell is None:
        raise ValueError('未在模板中找到“考核方式”段落。')
    footer_row = footer_cell.row

    table_start_row = header_cell.row
    table_end_row = footer_row - 1
    if table_end_row < table_start_row:
        raise ValueError('“日程安排”表格结构异常。')

    data_rows_per_block = table_end_row - data_start_row + 1
    if data_rows_per_block <= 0:
        raise ValueError('“日程安排”表格缺少数据行。')

    table_height = table_end_row - table_start_row + 1

    columns = {
        'week': 'B',
        'week_value': 'C',
        'period': 'D',
        'location': 'F',
        'teacher': 'G',
        'content': 'H',
    }

    table_merges = [
        (merge.min_row, merge.min_col, merge.max_row, merge.max_col)
        for merge in ws.merged_cells.ranges
        if table_start_row <= merge.min_row <= merge.max_row <= table_end_row
    ]

    column_indices = {header_cell.column}
    column_indices.update(column_index_from_string(letter) for letter in columns.values())
    column_indices.update(min_col for _min_row, min_col, _max_row, _max_col in table_merges)
    column_indices.update(max_col for _min_row, _min_col, _max_row, max_col in table_merges)
    col_start = min(column_indices)
    col_end = max(column_indices)

    template_rows = []
    for row in range(table_start_row, table_end_row + 1):
        row_cells = []
        for col in range(col_start, col_end + 1):
            cell = ws.cell(row=row, column=col)
            row_cells.append(
                {
                    'value': cell.value,
                    'has_style': cell.has_style,
                    'font': cell.font,
                    'fill': cell.fill,
                    'border': cell.border,
                    'alignment': cell.alignment,
                    'number_format': cell.number_format,
                    'protection': cell.protection,
                }
            )
        template_rows.append(row_cells)

    row_heights = {
        row: ws.row_dimensions[row].height
        for row in range(table_start_row, table_end_row + 1)
        if ws.row_dimensions[row].height is not None
    }

    if sessions:
        required_blocks = ceil(len(sessions) / data_rows_per_block)
    else:
        required_blocks = 1

    for block_index in range(1, required_blocks):
        insert_at = footer_row
        ws.insert_rows(insert_at, amount=table_height)

        for row_offset in range(table_height):
            template_row_index = table_start_row + row_offset
            target_row = insert_at + row_offset
            if template_row_index in row_heights:
                ws.row_dimensions[target_row].height = row_heights[template_row_index]

            for col_offset, template_cell in enumerate(template_rows[row_offset]):
                target_col = col_start + col_offset
                target_cell = ws.cell(row=target_row, column=target_col)
                target_cell.value = template_cell['value']
                if template_cell['has_style']:
                    if template_cell['font'] is not None:
                        target_cell.font = copy(template_cell['font'])
                    if template_cell['fill'] is not None:
                        target_cell.fill = copy(template_cell['fill'])
                    if template_cell['border'] is not None:
                        target_cell.border = copy(template_cell['border'])
                    if template_cell['alignment'] is not None:
                        target_cell.alignment = copy(template_cell['alignment'])
                    if template_cell['protection'] is not None:
                        target_cell.protection = copy(template_cell['protection'])
                target_cell.number_format = template_cell['number_format']

        row_shift = insert_at - table_start_row
        for min_row, min_col, max_row, max_col in table_merges:
            ws.merge_cells(
                start_row=min_row + row_shift,
                start_column=min_col,
                end_row=max_row + row_shift,
                end_column=max_col,
            )

        footer_row += table_height

    def safe_set(coord: str, value) -> None:
        cell = ws[coord]
        if isinstance(cell, MergedCell):
            return
        cell.value = value

    for block_index in range(required_blocks):
        block_data_start = data_start_row + block_index * table_height
        for row_idx in range(block_data_start, block_data_start + data_rows_per_block):
            safe_set(f"{columns['week']}{row_idx}", None)
            safe_set(f"{columns['week_value']}{row_idx}", None)
            safe_set(f"{columns['period']}{row_idx}", None)
            safe_set(f"{columns['location']}{row_idx}", None)
            safe_set(f"{columns['teacher']}{row_idx}", None)
            safe_set(f"{columns['content']}{row_idx}", None)

    for index, session in enumerate(sessions):
        block_index = index // data_rows_per_block
        row_offset = index % data_rows_per_block
        row_idx = data_start_row + block_index * table_height + row_offset
        safe_set(f"{columns['week']}{row_idx}", session.week_label)
        safe_set(f"{columns['week_value']}{row_idx}", '')  # 周值留空
        safe_set(f"{columns['period']}{row_idx}", session.period_label)
        safe_set(f"{columns['location']}{row_idx}", '')
        safe_set(f"{columns['teacher']}{row_idx}", session.teacher)
        safe_set(f"{columns['content']}{row_idx}", session.description)

    wb.save(output_path)


def generate_output(schedule_path: Path, template_path: Path, output_path: Path) -> Path:
    sessions = build_sessions(schedule_path)
    write_sessions_to_template(template_path, sessions, output_path)
    return output_path


def launch_gui_tkinter() -> bool:
    try:
        import tkinter as tk
        from tkinter import filedialog, messagebox
    except ImportError:
        return False

    try:
        root = tk.Tk()
    except tk.TclError as exc:
        print('无法初始化 tkinter 图形界面：', exc)
        return False
    root.title('实习日历填充工具')
    root.resizable(False, False)

    schedule_var = tk.StringVar(value=str(SCHEDULE_FILE.resolve()) if SCHEDULE_FILE.exists() else '')
    template_var = tk.StringVar(value=str(TEMPLATE_FILE.resolve()) if TEMPLATE_FILE.exists() else '')
    default_output = OUTPUT_FILE.resolve()
    output_var = tk.StringVar(value=str(default_output))

    def choose_schedule() -> None:
        path = filedialog.askopenfilename(
            title='选择课程表',
            filetypes=[('Excel 文件', '*.xlsx'), ('所有文件', '*.*')],
        )
        if path:
            schedule_var.set(path)

    def choose_template() -> None:
        path = filedialog.askopenfilename(
            title='选择模板',
            filetypes=[('Excel 文件', '*.xlsx'), ('所有文件', '*.*')],
        )
        if path:
            template_var.set(path)
            stem = Path(path).stem + '_填充.xlsx'
            output_var.set(str(Path(path).with_name(stem)))

    def choose_output() -> None:
        initial_file = Path(output_var.get()).name if output_var.get() else Path(template_var.get() or '输出.xlsx').stem + '.xlsx'
        path = filedialog.asksaveasfilename(
            title='选择导出位置',
            defaultextension='.xlsx',
            initialfile=initial_file,
            filetypes=[('Excel 文件', '*.xlsx')],
        )
        if path:
            output_var.set(path)

    status_var = tk.StringVar(value='全量导出将覆盖同名文件，请先备份。')

    def start_generation() -> None:
        schedule_path = Path(schedule_var.get()).expanduser()
        template_path = Path(template_var.get()).expanduser()
        output_path = Path(output_var.get()).expanduser()

        if not schedule_path.exists():
            messagebox.showerror('错误', '请先选择有效的课程表文件。')
            return
        if not template_path.exists():
            messagebox.showerror('错误', '请先选择有效的模板文件。')
            return

        try:
            status_var.set('正在处理，请稍候...')
            root.update_idletasks()
            generate_output(schedule_path, template_path, output_path)
        except Exception as exc:  # noqa: BLE001
            messagebox.showerror('生成失败', str(exc))
            status_var.set('生成失败，请检查输入文件。')
        else:
            messagebox.showinfo('成功', f'已生成文件：{output_path}')
            status_var.set(f'生成成功：{output_path}')

    padding = {'padx': 10, 'pady': 5}

    tk.Label(root, text='课程表：').grid(row=0, column=0, sticky='e', **padding)
    tk.Entry(root, textvariable=schedule_var, width=50).grid(row=0, column=1, **padding)
    tk.Button(root, text='选择...', command=choose_schedule).grid(row=0, column=2, **padding)

    tk.Label(root, text='模板：').grid(row=1, column=0, sticky='e', **padding)
    tk.Entry(root, textvariable=template_var, width=50).grid(row=1, column=1, **padding)
    tk.Button(root, text='选择...', command=choose_template).grid(row=1, column=2, **padding)

    tk.Label(root, text='导出路径：').grid(row=2, column=0, sticky='e', **padding)
    tk.Entry(root, textvariable=output_var, width=50).grid(row=2, column=1, **padding)
    tk.Button(root, text='浏览...', command=choose_output).grid(row=2, column=2, **padding)

    tk.Button(root, text='开始生成', command=start_generation, width=15).grid(row=3, column=1, **padding)
    tk.Label(root, textvariable=status_var, fg='gray').grid(row=4, column=0, columnspan=3, sticky='w', padx=10, pady=(0, 10))

    root.mainloop()
    return True


def launch_gui_pyside() -> bool:
    try:
        from PySide6.QtWidgets import (
            QApplication,
            QFileDialog,
            QGridLayout,
            QLabel,
            QLineEdit,
            QMessageBox,
            QPushButton,
            QWidget,
        )
    except ImportError as exc:
        print('未检测到 PySide6：', exc)
        print('当前解释器：', sys.executable)
        print('sys.path：', sys.path)
        return False

    try:
        app = QApplication.instance() or QApplication(sys.argv)

        class MainWindow(QWidget):  # pragma: no cover - GUI routine
            def __init__(self) -> None:
                super().__init__()
                self.setWindowTitle('实习日历填充工具 (Qt)')

                self.schedule_input = QLineEdit(str(SCHEDULE_FILE.resolve()) if SCHEDULE_FILE.exists() else '')
                self.template_input = QLineEdit(str(TEMPLATE_FILE.resolve()) if TEMPLATE_FILE.exists() else '')
                default_output = OUTPUT_FILE.resolve()
                self.output_input = QLineEdit(str(default_output))
                self.status_label = QLabel('全量导出将覆盖同名文件，请先备份。')

                layout = QGridLayout()
                layout.addWidget(QLabel('课程表：'), 0, 0)
                layout.addWidget(self.schedule_input, 0, 1)
                schedule_btn = QPushButton('选择...')
                schedule_btn.clicked.connect(self.choose_schedule)  # type: ignore[arg-type]
                layout.addWidget(schedule_btn, 0, 2)

                layout.addWidget(QLabel('模板：'), 1, 0)
                layout.addWidget(self.template_input, 1, 1)
                template_btn = QPushButton('选择...')
                template_btn.clicked.connect(self.choose_template)  # type: ignore[arg-type]
                layout.addWidget(template_btn, 1, 2)

                layout.addWidget(QLabel('导出路径：'), 2, 0)
                layout.addWidget(self.output_input, 2, 1)
                output_btn = QPushButton('浏览...')
                output_btn.clicked.connect(self.choose_output)  # type: ignore[arg-type]
                layout.addWidget(output_btn, 2, 2)

                generate_btn = QPushButton('开始生成')
                generate_btn.clicked.connect(self.start_generation)  # type: ignore[arg-type]
                layout.addWidget(generate_btn, 3, 1)

                layout.addWidget(self.status_label, 4, 0, 1, 3)

                self.setLayout(layout)

            def choose_schedule(self) -> None:
                path, _ = QFileDialog.getOpenFileName(self, '选择课程表', filter='Excel 文件 (*.xlsx);;所有文件 (*.*)')
                if path:
                    self.schedule_input.setText(path)

            def choose_template(self) -> None:
                path, _ = QFileDialog.getOpenFileName(self, '选择模板', filter='Excel 文件 (*.xlsx);;所有文件 (*.*)')
                if path:
                    self.template_input.setText(path)
                    output_name = Path(path).with_name(f"{Path(path).stem}_填充.xlsx")
                    self.output_input.setText(str(output_name))

            def choose_output(self) -> None:
                path, _ = QFileDialog.getSaveFileName(self, '选择导出位置', filter='Excel 文件 (*.xlsx)', selectedFilter='Excel 文件 (*.xlsx)')
                if path:
                    if not path.lower().endswith('.xlsx'):
                        path += '.xlsx'
                    self.output_input.setText(path)

            def start_generation(self) -> None:
                schedule_path = Path(self.schedule_input.text()).expanduser()
                template_path = Path(self.template_input.text()).expanduser()
                output_path = Path(self.output_input.text()).expanduser()

                if not schedule_path.exists():
                    QMessageBox.critical(self, '错误', '请先选择有效的课程表文件。')
                    return
                if not template_path.exists():
                    QMessageBox.critical(self, '错误', '请先选择有效的模板文件。')
                    return

                try:
                    self.status_label.setText('正在处理，请稍候...')
                    generate_output(schedule_path, template_path, output_path)
                except Exception as exc:  # noqa: BLE001
                    QMessageBox.critical(self, '生成失败', str(exc))
                    self.status_label.setText('生成失败，请检查输入文件。')
                else:
                    QMessageBox.information(self, '成功', f'已生成文件：{output_path}')
                    self.status_label.setText(f'生成成功：{output_path}')

        window = MainWindow()
        window.show()
        app.exec()
    except Exception as exc:  # noqa: BLE001
        print('无法初始化 PySide6 图形界面：', exc)
        return False
    return True


def launch_gui() -> None:
    if launch_gui_pyside():
        return
    if launch_gui_tkinter():
        return
    print('未检测到 tkinter 或 PySide6，无法启动图形界面。请安装任一库或使用命令行参数。')


def main(argv: List[str] | None = None) -> None:
    argv = argv if argv is not None else sys.argv[1:]

    if not argv:
        launch_gui()
        return

    parser = argparse.ArgumentParser(description='生成实习课程日历。')
    parser.add_argument('schedule', type=Path, help='课程表 Excel 文件路径')
    parser.add_argument('template', type=Path, help='模板 Excel 文件路径')
    parser.add_argument('output', type=Path, nargs='?', help='输出 Excel 文件路径')
    parsed = parser.parse_args(argv)

    schedule_path = parsed.schedule
    template_path = parsed.template
    output_path = parsed.output or template_path.with_name(f'{template_path.stem}_填充.xlsx')

    result = generate_output(schedule_path, template_path, output_path)
    print(f'已生成填充后的模板：{result}')


if __name__ == '__main__':
    main()

